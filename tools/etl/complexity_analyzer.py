"""
Analyze a CSV/Excel file and return complexity metadata.

Determines the best parse strategy ('pandas' or 'html') so downstream
importers can choose the right engine without trying both.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass
class ComplexityResult:
    is_complex: bool
    strategy: str   # 'pandas' | 'html'
    stats: dict[str, Any]


def analyze_complexity(path: str | Path) -> ComplexityResult:
    """Return parse strategy and file stats for *path*.

    Uses openpyxl metadata for .xlsx files; inspects raw bytes for CSV.
    """
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return ComplexityResult(
            is_complex=False,
            strategy="pandas",
            stats={"error": f"File not found: {path}"},
        )

    suffix = p.suffix.lower()
    file_size = p.stat().st_size

    if suffix in (".xlsx", ".xls", ".xlsm"):
        return _analyze_excel(p, file_size)

    return _analyze_csv(p, file_size)


def _analyze_excel(p: Path, file_size: int) -> ComplexityResult:
    try:
        import openpyxl  # type: ignore[import-not-found]
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb.active
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        wb.close()
        # Heuristic: merged cells or very many rows → 'html' strategy is safer
        is_complex = max_row > 50_000 or max_col > 30
        return ComplexityResult(
            is_complex=is_complex,
            strategy="html" if is_complex else "pandas",
            stats={
                "file_size_bytes": file_size,
                "sheets": sheet_names,
                "max_row": max_row,
                "max_col": max_col,
            },
        )
    except ImportError:
        return ComplexityResult(
            is_complex=False,
            strategy="pandas",
            stats={"file_size_bytes": file_size, "note": "openpyxl not installed"},
        )
    except Exception as exc:
        return ComplexityResult(
            is_complex=True,
            strategy="html",
            stats={"file_size_bytes": file_size, "error": str(exc)},
        )


def _analyze_csv(p: Path, file_size: int) -> ComplexityResult:
    raw = p.read_bytes()
    line_count = raw.count(b"\n")
    col_count = 0
    # Rough column count from first non-empty line
    for line in raw.splitlines():
        stripped = line.strip()
        if stripped:
            sep = b";" if stripped.count(b";") > stripped.count(b",") else b","
            col_count = len(stripped.split(sep))
            break

    is_complex = line_count > 50_000 or col_count > 30
    return ComplexityResult(
        is_complex=is_complex,
        strategy="pandas",  # CSV always uses pandas
        stats={
            "file_size_bytes": file_size,
            "approx_line_count": line_count,
            "approx_col_count": col_count,
        },
    )
